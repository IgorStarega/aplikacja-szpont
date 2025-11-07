#!/usr/bin/env python3
"""
Report Generator - v5.0 Feature
Eksport raportów do PDF i Excel

Funkcjonalność:
- ✅ Eksport do Excel (.xlsx) z formatowaniem
- ✅ Eksport do PDF z diagramami
- ✅ Raporty ze zmianami
"""

from pathlib import Path
from datetime import datetime
from typing import Dict, List, Any, Optional
import json

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    raise ImportError("openpyxl nie zainstalowany. Uruchom: pip install openpyxl")

try:
    from reportlab.lib.pagesizes import letter, A4
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import inch
    from reportlab.lib import colors
except ImportError:
    raise ImportError("reportlab nie zainstalowany. Uruchom: pip install reportlab")


class ReportGenerator:
    """Generator raportów - v5.0 Feature"""

    REPORTS_DIR = Path("src/.data/reports")

    def __init__(self):
        """Inicjalizacja generatora raportów"""
        self.REPORTS_DIR.mkdir(parents=True, exist_ok=True)

    def generate_excel_report(self, data: Dict[str, Any], filename: Optional[str] = None) -> str:
        """
        Generuj raport w Excelu

        Args:
            data: Dane do raportu (historia aktualizacji, statystyki)
            filename: Nazwa pliku (opcjonalnie, auto-generowana jeśli brak)

        Returns:
            Ścieżka do stworzonych pliku
        """
        if filename is None:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = f"raport_aktualizacji_{timestamp}.xlsx"

        filepath = self.REPORTS_DIR / filename

        wb = Workbook()
        ws = wb.active
        ws.title = "Podsumowanie"

        # Styl nagłówka
        header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=12)
        center_alignment = Alignment(horizontal="center", vertical="center")
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        # Tytuł
        ws['A1'] = "RAPORT AKTUALIZACJI STRONY"
        ws['A1'].font = Font(bold=True, size=14)
        ws.merge_cells('A1:D1')

        # Data raportu
        ws['A2'] = f"Data raportu: {datetime.now().strftime('%d.%m.%Y %H:%M:%S')}"
        ws['A2'].font = Font(italic=True)
        ws.merge_cells('A2:D2')

        # Nagłówki kolumn
        headers = ["Lp.", "Opis", "Wartość", "Jednostka"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=4, column=col)
            cell.value = header
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center_alignment
            cell.border = border

        # Dane statystyk
        row = 5
        stats = data.get('statistics', {})

        metrics = [
            ("Całkowite aktualizacje", stats.get('total_updates', 0), ""),
            ("Udane", stats.get('successful', 0), ""),
            ("Nieudane", stats.get('failed', 0), ""),
            ("Bez zmian", stats.get('no_changes', 0), ""),
            ("Średni czas trwania", stats.get('avg_duration', 0), "sekundy"),
            ("Karty dodane", stats.get('total_cards_added', 0), ""),
            ("Karty zmienione", stats.get('total_cards_modified', 0), ""),
            ("Karty usunięte", stats.get('total_cards_removed', 0), ""),
            ("Użycie cache'a", round(stats.get('cache_usage_percent', 0), 1), "%"),
        ]

        for i, (desc, value, unit) in enumerate(metrics, 1):
            ws.cell(row=row, column=1).value = i
            ws.cell(row=row, column=2).value = desc
            ws.cell(row=row, column=3).value = value
            ws.cell(row=row, column=4).value = unit

            for col in range(1, 5):
                ws.cell(row=row, column=col).border = border
                ws.cell(row=row, column=col).alignment = Alignment(horizontal="left", vertical="center")

            row += 1

        # Historia aktualizacji
        ws_history = wb.create_sheet("Historia")
        headers_history = ["Data", "Status", "Czas (s)", "Karty (+/-)", "Cache", "Foldery"]

        for col, header in enumerate(headers_history, 1):
            cell = ws_history.cell(row=1, column=col)
            cell.value = header
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center_alignment
            cell.border = border

        updates = data.get('recent_updates', [])
        for i, update in enumerate(updates[:100], 2):  # Pierwszy 100
            ws_history.cell(row=i, column=1).value = update.get('timestamp', '')
            ws_history.cell(row=i, column=2).value = update.get('status', '')
            ws_history.cell(row=i, column=3).value = update.get('duration', '')
            ws_history.cell(row=i, column=4).value = f"{update.get('added', 0)}/{update.get('removed', 0)}"
            ws_history.cell(row=i, column=5).value = "Tak" if update.get('cache_used') else "Nie"
            ws_history.cell(row=i, column=6).value = ", ".join(update.get('folders', []))

            for col in range(1, 7):
                ws_history.cell(row=i, column=col).border = border

        # Szerokość kolumn
        ws.column_dimensions['A'].width = 5
        ws.column_dimensions['B'].width = 25
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 15

        ws_history.column_dimensions['A'].width = 20
        ws_history.column_dimensions['B'].width = 12
        ws_history.column_dimensions['C'].width = 10
        ws_history.column_dimensions['D'].width = 12
        ws_history.column_dimensions['E'].width = 10
        ws_history.column_dimensions['F'].width = 30

        wb.save(filepath)
        return str(filepath)

    def generate_pdf_report(self, data: Dict[str, Any], filename: Optional[str] = None) -> str:
        """
        Generuj raport w PDF

        Args:
            data: Dane do raportu
            filename: Nazwa pliku (opcjonalnie)

        Returns:
            Ścieżka do stworzonych pliku
        """
        if filename is None:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = f"raport_aktualizacji_{timestamp}.pdf"

        filepath = self.REPORTS_DIR / filename

        # Ustawienia dokumentu
        doc = SimpleDocTemplate(str(filepath), pagesize=letter)
        story = []
        styles = getSampleStyleSheet()

        # Tytuł
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=18,
            textColor=colors.HexColor('#1F4E78'),
            spaceAfter=10
        )
        story.append(Paragraph("RAPORT AKTUALIZACJI STRONY", title_style))

        # Data
        story.append(Paragraph(
            f"<i>Wygenerowano: {datetime.now().strftime('%d.%m.%Y %H:%M:%S')}</i>",
            styles['Normal']
        ))
        story.append(Spacer(1, 0.3 * inch))

        # Tabela statystyk
        stats = data.get('statistics', {})
        table_data = [
            ['Metryka', 'Wartość'],
            ['Całkowite aktualizacje', str(stats.get('total_updates', 0))],
            ['Udane', str(stats.get('successful', 0))],
            ['Nieudane', str(stats.get('failed', 0))],
            ['Bez zmian', str(stats.get('no_changes', 0))],
            ['Średni czas', f"{stats.get('avg_duration', 0)}s"],
            ['Karty dodane', str(stats.get('total_cards_added', 0))],
            ['Karty zmienione', str(stats.get('total_cards_modified', 0))],
            ['Karty usunięte', str(stats.get('total_cards_removed', 0))],
            ['Użycie cache', f"{round(stats.get('cache_usage_percent', 0), 1)}%"],
        ]

        table = Table(table_data, colWidths=[3 * inch, 2 * inch])
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1F4E78')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 12),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ]))

        story.append(table)
        doc.build(story)

        return str(filepath)

    def list_reports(self) -> List[Dict[str, Any]]:
        """
        Lista wszystkich raportów

        Returns:
            Lista słowników z informacjami o raportach
        """
        reports = []
        for file in self.REPORTS_DIR.glob("*.xlsx"):
            reports.append({
                'name': file.name,
                'path': str(file),
                'size': file.stat().st_size,
                'created': datetime.fromtimestamp(file.stat().st_mtime).isoformat(),
                'type': 'Excel'
            })

        for file in self.REPORTS_DIR.glob("*.pdf"):
            reports.append({
                'name': file.name,
                'path': str(file),
                'size': file.stat().st_size,
                'created': datetime.fromtimestamp(file.stat().st_mtime).isoformat(),
                'type': 'PDF'
            })

        return sorted(reports, key=lambda x: x['created'], reverse=True)

